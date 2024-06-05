import openpyxl
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter

# Disable screen updating
# (No direct equivalent in Python, as it's not typically needed)
# Application.ScreenUpdating = False

# Create a new workbook
new_workbook = openpyxl.Workbook()
final_sheet = new_workbook.active
final_sheet.title = "FinalSheet"

# Load the source workbook
source_workbook = openpyxl.load_workbook("source_file.xlsx")

# Accessing individual sheets
ZCBB = source_workbook["ZCBB"]
ME2L = source_workbook["ME2L"]
oor = source_workbook["OOR"]
Transit = source_workbook["TRANSIT"]
MDA = source_workbook["MDA"]

# Find the last row and column with data
lastRow = ZCBB.max_row
lastCol = ZCBB.max_column
lastRowoor = oor.max_row
lastcoloor = oor.max_column
lastRowMDA = MDA.max_row
lastcolMDA = MDA.max_column
lastRowME2L = ME2L.max_row
lastcolME2L = ME2L.max_column
lastRowALV = Transit.max_row
lastcolALV = Transit.max_column

# Set destination starting range in the destination sheet
destination_range = final_sheet["A1"]

# Copy header from source sheet
for col in range(1, lastCol + 1):
    final_sheet.cell(row=1, column=col).value = ZCBB.cell(row=2, column=col).value

# Change the header names of additional columns
header_names = ["Purch_Doc", "Order Number", "Booked Date", "Promise Date", "Line/Shipment Number", "Ordered Quantity"]
for i, header_name in enumerate(header_names, start=lastCol + 1):
    final_sheet.cell(row=1, column=i).value = header_name

# Apply formatting to the new header rows
for col in range(lastCol + 1, lastCol + len(header_names) + 1):
    final_sheet.cell(row=1, column=col).fill = PatternFill(start_color="0000FF", end_color="0000FF", fill_type="solid")
    final_sheet.cell(row=1, column=col).font = Font(color="FFFFFF")

# Initialize variables
index = 2

# Loop through each row starting from row 3 to the last row
for currentRow in range(3, lastRow + 1):
    # Copy data from the current row up to the last column into the array
    dataArray = []
    for col in range(1, lastCol + 1):
        dataArray.append(ZCBB.cell(row=currentRow, column=col).value)

    # Paste the array into the specified destination in the destination sheet
    for col, value in enumerate(dataArray, start=1):
        final_sheet.cell(row=index, column=col).value = value

    # Reading pkey and matcode from source sheet
    sale_order = ZCBB.cell(row=currentRow, column=2).value
    order_Qty = ZCBB.cell(row=currentRow, column=26).value
    Mat_code = ZCBB.cell(row=currentRow, column=21).value

    # Applying filter on ME2M Sheet
    # (Assuming you have already filtered the data and you have the desired values)
    # Writing the filtered data to the destination sheet

    # Increment index for next row in destination sheet
    index += 1

# Save the new workbook
new_workbook.save("Final_Reports_ZCBB.xlsx")

# Enable screen updating
# (Not needed in Python)

# Optional: Close the source workbook
source_workbook.close()
